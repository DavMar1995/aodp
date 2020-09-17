import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

def initialize(file, wb, ws):
    global newFile
    if newFile:
        return
    else:
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        for row in range(ws.max_column):
            for cell in range(ws.max_row):
                # print(cell)
                # print(f'columns[{row}]cell+1 = {columns[row]}{cell+1}')
                id = f'{columns[row]}{cell+1}'
                print(f'id = {id}, type: {type(id)}')
                ws[id].value = None
        wb.save(file)

# corperations
def corperationInfo(delete = False):
    corperations = []
    wrWb.active = wrWb['corperation']
    wrWs = wrWb.active  
    
    if delete:
        initialize(wr, wrWb, wrWs)

    # add name and other info
    area = reWs['B1': 'N350']

    for b, c, d, e, f, g, h, i, j, k, l , m, n in area:
        if b.value == None:
            continue
        else:
            dict = {'name': b.value, 'capital': c.value, 'revenue': d.value, 'industry': e.value, 'field': f.value,
                    'url': j.value, 'address': k.value, 'telephone': l.value, 'fax': m.value, 'email': n.value}
            corperations.append(dict)

    # add the basic info
    area = reWs['J1': 'N350']

    # count = 0
    # for j, k, l, m, n in area:
    #     if j.value == None:
    #         continue
    #     else:
    #         corperations[count]['url'] = j.value
    #         corperations[count]['address'] = k.value
    #         corperations[count]['telephone'] = l.value
    #         corperations[count]['fax'] = m.value
    #         corperations[count]['email'] = n.value
    #         count += 1


    for i in range(len(corperations)):
        if i == 0:
            wrWs[f'A{i+1}'].value = '公司序號'
        else:
            wrWs[f'A{i+1}'].value = i
        wrWs[f'B{i+1}'].value = corperations[i]['name']
        wrWs[f'C{i+1}'].value = corperations[i]['capital']
        wrWs[f'D{i+1}'].value = corperations[i]['revenue']
        wrWs[f'E{i+1}'].value = corperations[i]['industry']
        wrWs[f'F{i+1}'].value = corperations[i]['field']
        wrWs[f'G{i+1}'].value = corperations[i]['url']
        wrWs[f'H{i+1}'].value = corperations[i]['address']
        wrWs[f'I{i+1}'].value = corperations[i]['telephone']
        wrWs[f'J{i+1}'].value = corperations[i]['fax']
        wrWs[f'K{i+1}'].value = corperations[i]['email']

    wrWb.save(wr)

# datasets
def datasetsInfo(delete = False):
    datasets = []
    wrWb.active = wrWb['dataset']
    wrWs = wrWb.active
    
    if delete:
        initialize(wr, wrWb, wrWs)
        
    area = reWs['G1': 'H350']
    for g, h in area:
        if g.value == None or h.value == None or '\n' in str(g.value):
            continue
        else:
            values = str(h.value).split('\n')
            for i in values:
                dict = {'name': i, 'provider': g.value}
                if dict not in datasets:
                    datasets.append(dict)

    for i in range(len(datasets)):
        if i == 0:            
            wrWs[f'A{i+1}'].value = '資料集序號'
            wrWs[f'B{i+1}'].value = '資料集名稱'
            wrWs[f'C{i+1}'].value = '提供部會'
        else:
            wrWs[f'A{i+1}'].value = i
            wrWs[f'B{i+1}'].value = datasets[i]['name']
            wrWs[f'C{i+1}'].value = datasets[i]['provider']
    wrWb.save(wr)


def dataInfo(sheetName, column, idName, delete = False, splitSymbol = '\n'):
    data = []
    
    wrWb.active = wrWb[sheetName]
    wrWs = wrWb.active

    if delete:
        initialize(wr, wrWb, wrWs)
    
    count = 0
    for cell in list(reWs.columns)[column]:
        if count == 0:
            data.append(cell.value)
        elif cell.value == None:
            continue
        else:
            values = str(cell.value).split(splitSymbol)
            for i in values:
                if i not in data:
                    data.append(i)
        count += 1
        
    for i in range(len(data)):
        wrWs[f'A{i+1}'] = i
        wrWs[f'B{i+1}'] = data[i]
    wrWs['A1'] = f'{sheetName}_Id'
        
    wrWb.save(wr)

def searchId(sheetName, comparedColumn, nameList):
    wrWb.active = wrWb[sheetName]
    wrWs = wrWb.active
    
    for value in nameList:
        if value in list(reWs.columns)[comparedColumn]:
            pass
    Id = 0
    # return Id
    pass

def datasetUsage():
    wrWb.active = wrWb['dataset_usage']
    wrWs = wrWb.active
    
    area = reWs['B3': 'H350']
    corName = ''
    sets = []
    for b, c, d, e, f,g, h in area:
        if not b.value ==  None:
            corName = b.value
        if h.value == None:
            continue
        else:
            values = str(h.value).split('\n')
            ids = []
            for i in values:
                if i not in sets:
                    #TODO: find the id of datasets in `dataset` sheet
                    sets.append(i)
            ids = searchId('dataset', 2, sets)
            for i, value in enumerate(ids):
                #TODO: write the usage of dataset of each corperation into `dataset_usage` table
                if i == 0:
                    wrWs[f'A{i+1}'].value = '流水號'
                    wrWs[f'B{i+1}'].value = '公司序號'
                    wrWs[f'C{i+1}'].value = '資料集序號'
                else:
                    wrWs[f'A{i+1}'].value = i
                    wrWs[f'B{i+1}'].value = corName
                    wrWs[f'C{i+1}'].value = value
    wrWb.save(wr)

def fieldsNormalization():
    
    pass

# def 


def printAll():
    for row in ws:
        print('------------------------------------\nrow start')
        for cell in row:
            print('-----\n')
            print(cell.value)
            print('\n-----')
        print('row end\n------------------------------------')


if __name__ == "__main__":

    wr = 'normalized_data.xlsx'
    re = 'OPEN100視覺化參數0401.xlsx'

    #! Choose one of two below options
    #* 1. create a new file, if the file is not existed currently
    # wrWb = openpyxl.Workbook()
    
    # wrWb.create_sheet('corperation', 0)
    # wrWb.create_sheet('industry', 1)
    # wrWb.create_sheet('field', 2)
    # wrWb.create_sheet('dataset', 3)
    # wrWb.create_sheet('ministry', 4)
    # wrWb.save(wr)
    
    # newFile = True

    #* 2. read the file, if it's existed
    wrWb = openpyxl.load_workbook(wr)
    wrWb.create_sheet('dataset_usage')
    # print(wrWb.active.title)
    newFile = False


    reWb = openpyxl.load_workbook(re)
    reWb.active = 0
    reWs = reWb.active 
       
    # ministriesInfo()
    corperationInfo(True) # update infomation of corperations
    datasetsInfo()
    
    dataInfo('ministry', 6, '部會序號')
    dataInfo('field', 5, '應用領域序號', splitSymbol=',')
    dataInfo('industry', 4, '行業序號')
        
    wrWb.save(wr)
    
    print('success')