from bs4 import BeautifulSoup
import requests
import pandas as pd
from post import getHtml
import openpyxl


def crawlData(code, func, year, month, kind):
    response = getHtml(code, func, year, month, kind)

    soup = BeautifulSoup(response.text, "html.parser")

    # print(soup)

    div = soup.find("div", id="content")
    table = div.select_one("table")

    return getData(table, year, month)


def getData(table, year, month):
    data = []
    head = []
    summary = []
    # finished = False

    ths = table.select("th")
    for th in ths:
        head.append(th.string)
    head.append("Time")

    trs = table.select("tbody tr")
    for tr in trs:
        tds = tr.select("td")
        dTmp = []
        sTmp = []
        for td in tds:
            text = td.getText()
            # if text == "Total":
            #     finished = True
            # if finished:
            #     # if text == "" or text == " ":
            #     #     break
            #     # else:
            #     sTmp.append(text)
            # else:
            dTmp.append(text)
        dTmp.append(f'{year}-{month}')

        # if finished:
        #     summary.append(sTmp)
        # else:
        data.append(dTmp)

    # print(trs.find_all("td"))

    dDf = pd.DataFrame(data, columns=head)
    # print(dDf)

    sDf = pd.DataFrame(summary, columns=head)
    # print(sDf)

    dDf.to_excel(f'{year}-{month}.xlsx', index=False)
    dDf.to_excel('test.xlsx', index=False)
    return dDf


def getFileName():
    # year = div.select_one("input", id="yyyy")
    # month = div.select_one("input", id="mm")
    year = div.find("select", id="selectYear").find(
        "option", selected=True).getText()
    month = div.find("select", id="selectMonth").find(
        "option", selected=True).getText()
    #! I can't deal with it.
    # dataTypes = soup.findAll("input", attrs={'value':"1", 'type':"radio", 'name':"radioMonth", 'checked':True})
    # for dataType in dataTypes:
    #     dataType.find("input", checked=True)
    #     print(dataType)
    # print(dataTypes)

    print(year, month)
    # print(year.getText())

# move all the cells down for  one row


def moveCells():
    fn = 'test.xlsx'
    wb = openpyxl.load_workbook(fn)

    wb.active = 0
    ws = wb.active

    odd = []

    columns = [
        'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
        'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
        'U', 'V', 'W', 'X', 'Y', 'Z']

    for row in range(1, ws.max_row):
        tmp = []

        # if row == 1:
        # continue
        print(f'row = {row}, {ws[f"A{row}"].value}')
        print(odd)
        for cell in range(1, ws.max_column):
            print(f'\ncell = {cell}')
            # print(f'columns[{row}]cell+1 = {columns[row]}{cell+1}')
            id = f'{columns[cell-1]}{row}'
            # print(f'id = {id}, type: {type(id)}')
            tmp.append(ws[id].value)
            print(f'id = {id}, row = {row}, cell = {cell}')
            if row == 1:
                # ws[id] =
                continue
            ws[id].value = odd[row-2][cell-1]
            print(odd)

        odd.append(tmp)
        # print(odd)

    for i in range(1, ws.max_row):
        a = []
        for j in range(1, ws.max_column):
            a.append(ws[f'{columns[j]}{i+1}'].value)
        print(a)

    wb.save(fn)

    # for row in range():
    # pass
if __name__ == '__main__':
    # response = requests.get(
    #     "https://kto.visitkorea.or.kr/eng/tourismStatics/keyFacts/KoreaMonthlyStatistics/eng/inout/inout.kto")

    YYYY = '2020'
    MM = '06'

    CODE = '21'
    # * Gender & Nationality: 21
    # * Purpose & Nationality: 22
    # * Age & Nationality: 23
    # * Transport & Nationality: 25 // ERROR

    FUNC = '1'
    # * Statistics of Arrivals & Departures by Item: 1, 5, 0
    # * Statistics of Tourism Receipt & Expenditure: 2
    # * Statistics Arrivals & Departures by year: 3, 4
    # * Statistics of Arrivals by Cruise: ?

    TYPE = '1'
    # * By Month: 1
    # * Accumulative Data: 2

    dataFrameDict = []

    for y in range(int(YYYY), 2021):
        if y != 2020:
            for m in range(1, 13):
                if m < 10:
                    m = '0'+str(m)
                else:
                    m = str(m)
                print(f'\n\nTHIS IS {y}-{m}\n')
                dataFrameDict.append(crawlData(
                    CODE, FUNC, str(y), str(m), TYPE))
        else:
            for m in range(1, int(MM)+1):
                if m < 10:
                    m = '0'+str(m)
                else:
                    m = str(m)
                print(f'\n\nTHIS IS {y}-{m}\n')
                dataFrameDict.append(crawlData(CODE, FUNC, str(y), m, TYPE))

    df = pd.concat(dataFrameDict, ignore_index=True, axis=0)
    # print(df)

    # nwb = openpyxl.load_workbook(fn)
    df.to_excel(f'{YYYY}_2020-{MM}.xlsx', index=False)

    # moveCells()
